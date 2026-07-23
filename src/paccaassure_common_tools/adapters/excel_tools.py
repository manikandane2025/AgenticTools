from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
import xlsxwriter

from paccaassure_common_tools.adapters.common import (
    build_source_artifact,
    canonical_checksum,
    evidence,
    finalize_result,
    library_versions,
    resolve_input_file,
    rows_to_canonical_table,
    warning,
)
from paccaassure_common_tools.artifacts import ArtifactCollector
from paccaassure_common_tools.exceptions import InputValidationError
from paccaassure_common_tools.interfaces import ToolAdapter
from paccaassure_common_tools.models import (
    CanonicalTableOutput,
    CertificationVerdict,
    NetworkPolicy,
    StagedArtifact,
    ToolCapability,
    ToolIdentity,
    ToolInvocationContext,
    ToolMaturity,
    ToolMetrics,
    ToolRegistration,
    ToolResult,
)
from paccaassure_common_tools.version import PACKAGE_VERSION


def _load_workbook(path: Path, *, data_only: bool = False) -> openpyxl.Workbook:
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=data_only)
    except Exception as exc:  # noqa: BLE001
        raise InputValidationError(
            "The workbook could not be opened.",
            details={"file_name": path.name, "exception_type": type(exc).__name__},
        ) from exc


def _iter_sheet_rows(sheet: openpyxl.worksheet.worksheet.Worksheet, values_only: bool = True) -> list[list[object]]:
    return [list(row) for row in sheet.iter_rows(values_only=values_only)]


def _sheet_headers(
    rows: list[list[object]], header_row: int | None
) -> tuple[list[str], list[list[object]], int]:
    if not rows:
        return [], [], 1
    index = (header_row - 1) if header_row else 0
    headers = ["" if value is None else str(value) for value in rows[index]]
    data_rows = rows[index + 1 :]
    return headers, data_rows, index + 2


def _optional_int(value: object) -> int | None:
    return value if isinstance(value, int) else None


def _list_of_strings(value: object) -> list[str]:
    if not isinstance(value, list):
        return []
    return [str(item) for item in value]


def _source(path: Path) -> dict[str, object]:
    return build_source_artifact(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        detected_format=path.suffix,
    )


class ExcelInspectTool(ToolAdapter):
    def execute(
        self,
        payload: dict[str, object],
        context: ToolInvocationContext,
        collector: ArtifactCollector,
    ) -> ToolResult:
        del collector
        path = resolve_input_file(payload, context)
        packages = library_versions("openpyxl", "XlsxWriter")
        workbook = _load_workbook(path)
        source = _source(path)
        sheets = []
        formula_count = 0
        hidden_sheets = 0
        for sheet in workbook.worksheets:
            sample = _iter_sheet_rows(sheet)[:5]
            headers = (
                [str(cell) if cell is not None else "" for cell in sample[0]] if sample else []
            )
            formula_count += sum(
                1
                for row in sample
                for cell in row
                if isinstance(cell, str) and cell.startswith("=")
            )
            if sheet.sheet_state != "visible":
                hidden_sheets += 1
            sheets.append(
                {
                    "name": sheet.title,
                    "visible": sheet.sheet_state == "visible",
                    "dimensions": sheet.calculate_dimension(),
                    "header_candidates": headers,
                    "sample_rows": sample[1:4] if len(sample) > 1 else [],
                }
            )
        warnings = []
        if path.suffix == ".xlsm":
            warnings.append(
                warning(
                    "EXCEL_MACRO_CONTENT_IGNORED",
                    "Macro-enabled workbook loaded in read mode.",
                    source_scope={"file_name": path.name},
                    policy_reference="macro_execution",
                )
            )
        outputs = {
            "workbook": {
                "file_name": path.name,
                "sheet_count": len(sheets),
                "active_sheet": (
                    workbook.active.title
                    if workbook.worksheets and workbook.active is not None
                    else None
                ),
                "sheets": sheets,
                "formula_count": formula_count,
            }
        }
        evidence_items = [
            evidence(
                context=context,
                kind="excel_inspect",
                source_checksum=str(source["source_checksum"]),
                output_checksum=canonical_checksum(outputs),
                capability_ids=["excel.inspect_workbook_structure"],
                details={"sheet_count": len(sheets), "formula_count": formula_count},
                source_artifact_ref=path.name,
                fixture_identity=path.name,
            )
        ]
        return finalize_result(
            context=context,
            packages=packages,
            outputs=outputs,
            warnings=warnings,
            evidence=evidence_items,
            metrics=ToolMetrics(
                input_bytes=path.stat().st_size,
                sheets_discovered=len(sheets),
                sheets_processed=len(sheets),
                formulas_encountered=formula_count,
                hidden_sheets_skipped=hidden_sheets,
                adapter_library_versions=packages,
            ),
            capabilities_exercised=["excel.inspect_workbook_structure"],
            source_artifacts=[source],
            selection={"mode": "inspect"},
            policies={"macro_execution": "deny"},
        )


class ExcelReadTool(ToolAdapter):
    def execute(
        self,
        payload: dict[str, object],
        context: ToolInvocationContext,
        collector: ArtifactCollector,
    ) -> ToolResult:
        del collector
        path = resolve_input_file(payload, context)
        packages = library_versions("openpyxl")
        workbook = _load_workbook(path, data_only=False)
        source = _source(path)
        selected_sheets = set(_list_of_strings(payload.get("sheets"))) or None
        header_row = _optional_int(payload.get("header_row"))
        tables = []
        total_rows = 0
        rows_discovered = 0
        for sheet in workbook.worksheets:
            if selected_sheets and sheet.title not in selected_sheets:
                continue
            rows = _iter_sheet_rows(sheet)
            rows_discovered += len(rows)
            headers, data_rows, row_offset = _sheet_headers(rows, header_row)
            if rows and not headers:
                raise InputValidationError(
                    "Failed to parse headers.", details={"sheet": sheet.title}
                )
            def row_source(row_number: int, width: int, *, sheet_name: str = sheet.title) -> dict[str, object]:
                return {
                    "sheet_name": sheet_name,
                    "header_row": header_row or 1,
                    "cell_range": f"A{row_number}:{openpyxl.utils.get_column_letter(max(width, 1))}{row_number}",
                    "hidden": False,
                    "formula_mode": "expression",
                    "workbook_checksum": source["source_checksum"],
                }
            tables.append(
                rows_to_canonical_table(
                    name=sheet.title,
                    source={"type": "excel", "file_name": path.name, "sheet_name": sheet.title},
                    headers=headers,
                    rows=data_rows,
                    row_offset=row_offset,
                    source_builder=row_source,
                )
            )
            total_rows += len(data_rows)
        if not tables:
            raise InputValidationError("No sheets were selected.", details={"file_name": path.name})
        output = CanonicalTableOutput(
            tables=tables,
            metrics={"sheet_count": len(tables), "row_count": total_rows},
            provenance={"file_name": path.name},
        )
        outputs = {"table_output": output.model_dump(mode="json")}
        evidence_items = [
            evidence(
                context=context,
                kind="excel_read",
                source_checksum=str(source["source_checksum"]),
                output_checksum=canonical_checksum(outputs),
                capability_ids=["excel.read_canonical_tables"],
                details={"sheet_count": len(tables), "row_count": total_rows},
                source_artifact_ref=path.name,
                fixture_identity=path.name,
            )
        ]
        return finalize_result(
            context=context,
            packages=packages,
            outputs=outputs,
            warnings=[],
            evidence=evidence_items,
            metrics=ToolMetrics(
                input_bytes=path.stat().st_size,
                sheets_processed=len(tables),
                rows_discovered=rows_discovered,
                rows_returned=total_rows,
                records_processed=total_rows,
                records_returned=total_rows,
                tables_returned=len(tables),
                header_rows_consumed=1,
                adapter_library_versions=packages,
            ),
            capabilities_exercised=["excel.read_canonical_tables"],
            source_artifacts=[source],
            selection={"selected_sheets": sorted(selected_sheets) if selected_sheets else "all"},
            policies={"formula_mode": "expression"},
        )


class ExcelValidateTool(ToolAdapter):
    def execute(
        self,
        payload: dict[str, object],
        context: ToolInvocationContext,
        collector: ArtifactCollector,
    ) -> ToolResult:
        del collector
        path = resolve_input_file(payload, context)
        packages = library_versions("openpyxl")
        workbook = _load_workbook(path)
        source = _source(path)
        required_headers = _list_of_strings(payload.get("required_headers"))
        sheet_name = str(payload.get("sheet", workbook.sheetnames[0] if workbook.sheetnames else ""))
        if sheet_name not in workbook.sheetnames:
            raise InputValidationError("Required sheet is missing.", details={"sheet": sheet_name})
        rows = _iter_sheet_rows(workbook[sheet_name])
        headers, data_rows, _ = _sheet_headers(rows, _optional_int(payload.get("header_row")))
        missing_headers = [header for header in required_headers if header not in headers]
        duplicate_headers = sorted({header for header in headers if headers.count(header) > 1 and header})
        warnings = []
        if path.suffix == ".xlsm":
            warnings.append(
                warning(
                    "EXCEL_MACRO_CONTENT_IGNORED",
                    "Macro-enabled workbook content was validated without executing macros.",
                    source_scope={"file_name": path.name},
                    policy_reference="macro_execution",
                )
            )
        outputs = {
            "valid": not missing_headers and not duplicate_headers,
            "sheet": sheet_name,
            "row_count": len(data_rows),
            "headers": headers,
            "missing_headers": missing_headers,
            "duplicate_headers": duplicate_headers,
        }
        evidence_items = [
            evidence(
                context=context,
                kind="excel_validate",
                source_checksum=str(source["source_checksum"]),
                output_checksum=canonical_checksum(outputs),
                capability_ids=["excel.validate_structure_and_headers"],
                details=outputs,
                source_artifact_ref=path.name,
                fixture_identity=path.name,
            )
        ]
        return finalize_result(
            context=context,
            packages=packages,
            outputs=outputs,
            warnings=warnings,
            evidence=evidence_items,
            metrics=ToolMetrics(
                input_bytes=path.stat().st_size,
                sheets_processed=1,
                records_processed=len(data_rows),
                rows_returned=len(data_rows),
                header_rows_consumed=1,
                adapter_library_versions=packages,
            ),
            capabilities_exercised=["excel.validate_structure_and_headers"],
            source_artifacts=[source],
            selection={"sheet": sheet_name},
            policies={"macro_execution": "deny"},
        )


class ExcelWriteTool(ToolAdapter):
    def execute(
        self,
        payload: dict[str, object],
        context: ToolInvocationContext,
        collector: ArtifactCollector,
    ) -> ToolResult:
        sheets = payload.get("sheets")
        if not isinstance(sheets, list) or not sheets:
            raise InputValidationError("Workbook write requires sheets.", details={"field": "sheets"})
        packages = library_versions("XlsxWriter")
        stage = collector.stage_path("excel_write_output.xlsx")
        workbook = xlsxwriter.Workbook(str(stage))
        try:
            for sheet_payload in sheets:
                if not isinstance(sheet_payload, dict):
                    raise InputValidationError("Invalid sheet payload.", details={"sheet_payload": "not-an-object"})
                sheet = workbook.add_worksheet(str(sheet_payload.get("name", "Sheet1"))[:31])
                headers = sheet_payload.get("headers", [])
                rows = sheet_payload.get("rows", [])
                for col_index, header in enumerate(headers):
                    sheet.write(0, col_index, header)
                for row_index, row in enumerate(rows, start=1):
                    for col_index, value in enumerate(row):
                        sheet.write(row_index, col_index, value)
                if sheet_payload.get("freeze_panes"):
                    sheet.freeze_panes(1, 0)
            workbook.close()
        except Exception:
            workbook.close()
            raise
        artifact = collector.commit(
            StagedArtifact(
                temp_path=stage,
                final_name="excel_write_output.xlsx",
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        )
        evidence_items = [
            evidence(
                context=context,
                kind="excel_write",
                source_checksum=None,
                output_checksum=artifact.sha256,
                capability_ids=["excel.write_new_workbook_artifact"],
                details={"artifact_name": artifact.logical_name, "sheet_count": len(sheets)},
                artifact_refs=[artifact.logical_name],
            )
        ]
        artifact.evidence_ref = evidence_items[0].evidence_id
        return finalize_result(
            context=context,
            packages=packages,
            outputs={"artifact_path": artifact.path},
            warnings=[],
            evidence=evidence_items,
            metrics=ToolMetrics(
                sheets_processed=len(sheets),
                output_bytes_written=artifact.size_bytes,
                artifacts_created=1,
                adapter_library_versions=packages,
            ),
            capabilities_exercised=["excel.write_new_workbook_artifact"],
            policies={"write_mode": "new_workbook_only"},
            artifacts=[artifact],
        )


class ExcelCompareTool(ToolAdapter):
    def execute(
        self,
        payload: dict[str, object],
        context: ToolInvocationContext,
        collector: ArtifactCollector,
    ) -> ToolResult:
        del collector
        left_rel = payload.get("left_path")
        right_rel = payload.get("right_path")
        if not isinstance(left_rel, str) or not isinstance(right_rel, str):
            raise InputValidationError("Workbook compare requires both paths.")
        left = resolve_input_file({"path": left_rel}, context)
        right = resolve_input_file({"path": right_rel}, context)
        packages = library_versions("openpyxl")
        left_book = _load_workbook(left, data_only=False)
        right_book = _load_workbook(right, data_only=False)
        diffs: list[dict[str, object]] = []
        for sheet_name in sorted(set(left_book.sheetnames) | set(right_book.sheetnames)):
            if sheet_name not in left_book.sheetnames or sheet_name not in right_book.sheetnames:
                diffs.append({"sheet": sheet_name, "difference": "missing_sheet"})
                continue
            left_rows = _iter_sheet_rows(left_book[sheet_name])
            right_rows = _iter_sheet_rows(right_book[sheet_name])
            max_len = max(len(left_rows), len(right_rows))
            for index in range(max_len):
                lrow = left_rows[index] if index < len(left_rows) else []
                rrow = right_rows[index] if index < len(right_rows) else []
                if lrow != rrow:
                    diffs.append({"sheet": sheet_name, "row_number": index + 1, "left": lrow, "right": rrow})
        outputs = {"matches": not diffs, "differences": diffs}
        evidence_items = [
            evidence(
                context=context,
                kind="excel_compare",
                source_checksum=canonical_checksum([_source(left), _source(right)]),
                output_checksum=canonical_checksum(outputs),
                capability_ids=["excel.compare_sheet_rows"],
                details={"difference_count": len(diffs)},
                fixture_identity=f"{left.name}|{right.name}",
            )
        ]
        return finalize_result(
            context=context,
            packages=packages,
            outputs=outputs,
            warnings=[],
            evidence=evidence_items,
            metrics=ToolMetrics(
                input_bytes=left.stat().st_size + right.stat().st_size,
                records_processed=len(diffs),
                records_returned=len(diffs),
                adapter_library_versions=packages,
            ),
            capabilities_exercised=["excel.compare_sheet_rows"],
            source_artifacts=[_source(left), _source(right)],
            policies={"comparison_mode": "row_oriented"},
        )


def register_excel_tools(registry: Any) -> None:
    registrations: list[tuple[str, ToolAdapter, list[ToolCapability]]] = [
        (
            "excel_inspect",
            ExcelInspectTool(),
            [
                ToolCapability(
                    name="excel.inspect_workbook_structure",
                    supported_formats=[".xlsx", ".xlsm"],
                    supported_modes=["inspect"],
                    limits={"max_file_size_bytes": 50_000_000, "max_sheet_count": 100},
                    deterministic=True,
                    network_requirement=NetworkPolicy.DENY,
                    known_fidelity_restrictions=["No .xls support", "No macro execution"],
                )
            ],
        ),
        (
            "excel_read",
            ExcelReadTool(),
            [
                ToolCapability(
                    name="excel.read_canonical_tables",
                    supported_formats=[".xlsx", ".xlsm"],
                    supported_modes=["read"],
                    limits={"max_file_size_bytes": 50_000_000, "max_sheet_count": 100},
                    deterministic=True,
                    network_requirement=NetworkPolicy.DENY,
                    known_fidelity_restrictions=[
                        "Header detection defaults to first row when header_row is omitted",
                        "No explicit cell-range selection in v1",
                    ],
                )
            ],
        ),
        (
            "excel_validate",
            ExcelValidateTool(),
            [
                ToolCapability(
                    name="excel.validate_structure_and_headers",
                    supported_formats=[".xlsx", ".xlsm"],
                    supported_modes=["validate"],
                    limits={"max_file_size_bytes": 50_000_000, "max_sheet_count": 100},
                    deterministic=True,
                    network_requirement=NetworkPolicy.DENY,
                    known_fidelity_restrictions=["Schema validation is limited to required headers and duplicate-header checks"],
                )
            ],
        ),
        (
            "excel_write",
            ExcelWriteTool(),
            [
                ToolCapability(
                    name="excel.write_new_workbook_artifact",
                    supported_formats=[".xlsx"],
                    supported_modes=["write"],
                    limits={"max_output_bytes": 50_000_000, "max_sheet_count": 100},
                    deterministic=True,
                    network_requirement=NetworkPolicy.DENY,
                    known_fidelity_restrictions=["Write support is limited to new workbook creation", "Styles are write-limited"],
                )
            ],
        ),
        (
            "excel_compare",
            ExcelCompareTool(),
            [
                ToolCapability(
                    name="excel.compare_sheet_rows",
                    supported_formats=[".xlsx", ".xlsm"],
                    supported_modes=["compare"],
                    limits={"max_file_size_bytes": 50_000_000, "max_sheet_count": 100},
                    deterministic=True,
                    network_requirement=NetworkPolicy.DENY,
                    known_fidelity_restrictions=["Diff output is row-oriented and does not include style deltas"],
                )
            ],
        ),
    ]
    for tool_key, adapter, capabilities in registrations:
        registry.register(
            ToolRegistration(
                identity=ToolIdentity(
                    tool_key=tool_key,
                    version=PACKAGE_VERSION,
                    family="excel",
                    adapter_key=f"pacca_tools.io.{tool_key}",
                ),
                capabilities=capabilities,
                maturity=ToolMaturity.CERTIFIED,
                certification=CertificationVerdict.CERTIFIED,
                adapter=adapter,
            )
        )
