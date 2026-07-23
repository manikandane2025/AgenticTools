from __future__ import annotations

import csv
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo
from pypdf import PdfReader, PdfWriter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, TableStyle
from reportlab.platypus import Table as PdfTable

ROOT = Path(__file__).resolve().parents[1] / "tests" / "fixtures"


def create_excel() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["ID", "Name", "Value", "Created", "Formula", "Link"])
    for index in range(1, 121):
        sheet.append(
            [
                f"A-{index:03d}",
                f"Row {index}",
                index,
                date(2026, 7, min(index, 28)),
                f"=C{index + 1}*2",
                f"https://example.com/{index}",
            ]
        )
    sheet["B2"].comment = Comment("first row comment", "codex")
    sheet["F2"].hyperlink = "https://example.com/1"
    sheet["A1"].font = Font(bold=True)
    sheet.merge_cells("H1:I1")
    sheet["H1"] = "MergedHeader"
    table = Table(displayName="MainTable", ref="A1:F121")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    workbook.defined_names["DataStart"] = DefinedName("DataStart", attr_text="'Sheet1'!$A$2")
    extra = workbook.create_sheet("HiddenSheet")
    extra.sheet_state = "hidden"
    extra.append(["Flag", "Status"])
    extra.append(["Y", "hidden"])
    workbook.save(ROOT / "normal_workbook.xlsx")

    dup = openpyxl.Workbook()
    dsheet = dup.active
    dsheet.append(["ID", "ID", "Value"])
    dsheet.append(["1", "duplicate", "x"])
    dup.save(ROOT / "duplicate_headers.xlsx")

    left = openpyxl.Workbook()
    lsheet = left.active
    lsheet.title = "Compare"
    lsheet.append(["ID", "Value"])
    lsheet.append(["1", 10])
    lsheet.append(["2", 20])
    left.save(ROOT / "compare_left.xlsx")

    right = openpyxl.Workbook()
    rsheet = right.active
    rsheet.title = "Compare"
    rsheet.append(["ID", "Value"])
    rsheet.append(["1", 10])
    rsheet.append(["2", 30])
    rsheet.append(["3", 40])
    right.save(ROOT / "compare_right.xlsx")

    empty = openpyxl.Workbook()
    empty.active.title = "Empty"
    empty.save(ROOT / "empty_workbook.xlsx")

    (ROOT / "corrupt_workbook.xlsx").write_bytes(b"not-a-real-workbook")


def create_csv() -> None:
    rows = [["id", "name", "value"], ["1", "alpha", "100"], ["2", "beta", "200"]]
    with (ROOT / "comma_utf8.csv").open("w", encoding="utf-8", newline="") as handle:
        csv.writer(handle).writerows(rows)
    with (ROOT / "semicolon_utf8.csv").open("w", encoding="utf-8", newline="") as handle:
        csv.writer(handle, delimiter=";").writerows(rows)
    with (ROOT / "tab_utf8.tsv").open("w", encoding="utf-8", newline="") as handle:
        csv.writer(handle, delimiter="\t").writerows(rows)
    with (ROOT / "pipe_utf8.csv").open("w", encoding="utf-8", newline="") as handle:
        csv.writer(handle, delimiter="|").writerows(rows)
    with (ROOT / "quoted_newline.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["id", "name", "value"])
        writer.writerow(["1", "alpha,inc", "line1\nline2"])
    with (ROOT / "utf16.csv").open("w", encoding="utf-16", newline="") as handle:
        csv.writer(handle).writerows(rows)
    with (ROOT / "no_header.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["1", "alpha", "100"])
        writer.writerow(["2", "beta", "200"])
    with (ROOT / "malformed.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["id", "name", "value"])
        writer.writerow(["1", "broken"])
    (ROOT / "empty.csv").write_text("", encoding="utf-8")


def create_text_pdf(path: Path, text: str, *, page_count: int = 1) -> None:
    from reportlab.pdfgen import canvas

    pdf = canvas.Canvas(str(path), pagesize=letter)
    for index in range(page_count):
        pdf.drawString(72, 720, f"{text} page {index + 1}")
        if index + 1 < page_count:
            pdf.showPage()
    pdf.save()


def create_blank_pdf(path: Path, *, page_count: int = 1) -> None:
    from reportlab.pdfgen import canvas

    pdf = canvas.Canvas(str(path), pagesize=letter)
    for index in range(page_count):
        pdf.rect(72, 640, 64, 64, stroke=1, fill=0)
        if index + 1 < page_count:
            pdf.showPage()
    pdf.save()


def create_table_pdf(path: Path) -> None:
    document = SimpleDocTemplate(str(path), pagesize=letter)
    data = [["ID", "NAME", "VALUE"], ["1", "Alpha", "10"], ["2", "Beta", "20"]]
    table = PdfTable(data, colWidths=[100, 160, 100])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]
        )
    )
    document.build([table])


def create_rotated_pdf(path: Path) -> None:
    source = ROOT / "text.pdf"
    reader = PdfReader(str(source))
    writer = PdfWriter()
    page = reader.pages[0]
    page.rotate(90)
    writer.add_page(page)
    with path.open("wb") as handle:
        writer.write(handle)


def create_encrypted_pdf(path: Path) -> None:
    source = ROOT / "text.pdf"
    reader = PdfReader(str(source))
    writer = PdfWriter()
    for page in reader.pages:
        writer.add_page(page)
    writer.encrypt("secret")
    with path.open("wb") as handle:
        writer.write(handle)


def create_mixed_pdf(path: Path) -> None:
    text_reader = PdfReader(str(ROOT / "text.pdf"))
    blank_reader = PdfReader(str(ROOT / "blank.pdf"))
    writer = PdfWriter()
    writer.add_page(text_reader.pages[0])
    writer.add_page(blank_reader.pages[0])
    with path.open("wb") as handle:
        writer.write(handle)


def create_corrupt_pdf(path: Path) -> None:
    path.write_bytes(b"%PDF-1.4\ncorrupt")


def create_fixtures() -> None:
    ROOT.mkdir(parents=True, exist_ok=True)
    create_excel()
    create_csv()
    create_text_pdf(ROOT / "text.pdf", "PaccaAssure PDF text fixture")
    create_text_pdf(ROOT / "multi_page.pdf", "PaccaAssure multi page fixture", page_count=2)
    create_blank_pdf(ROOT / "blank.pdf")
    create_table_pdf(ROOT / "table.pdf")
    create_rotated_pdf(ROOT / "rotated.pdf")
    create_encrypted_pdf(ROOT / "encrypted.pdf")
    create_mixed_pdf(ROOT / "mixed.pdf")
    create_corrupt_pdf(ROOT / "corrupt.pdf")


if __name__ == "__main__":
    create_fixtures()
